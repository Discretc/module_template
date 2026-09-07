import sqlite3
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

import database  # noqa: E402
import import_excel  # noqa: E402
from rules import JointRuleConflictError, RuleValidationError  # noqa: E402


MASTER_HEADERS = [
    "Faculty_Code", "Faculty _Chn", "Faculty _Eng", "Faculty _Prt",
    "Prog_Code", "Prog_Chn", "Prog_Eng", "Prog_Prt", "Class_Code",
    "Module_Chn", "Module_Eng", "Module_Prt", "Prerequsite_Chn",
    "Prerequsite_Eng", "Prerequsite_Por", "Credits", "Durations",
    "Instructor_Chn", "Instructor_Eng", "Instructor_Prt", "Email",
    "Room_Chn", "Room_Eng", "Room_Prt", "Telephone", "Rule",
    "Joint_Relationship",
]


def master_row(programme_code, programme_name, class_code, related="", rule=None, prerequisite=None):
    return [
        "FCA", "應用科學學院", "Faculty of Applied Sciences", "Faculdade de Ciências Aplicadas",
        programme_code, f"{programme_name}中", programme_name, f"{programme_name} PT", class_code,
        "模組", "Module", "Módulo", prerequisite, prerequisite, prerequisite,
        3, 45, "教師", "LECTURER", "DOCENTE", "lecturer@mpu.edu.mo",
        None, None, None, 85990000, rule, related,
    ]


def write_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "工作表1"
    sheet.append(MASTER_HEADERS)
    sheet.append(master_row("P1", "Bachelor One", "COMP1000-111", "COMP1000-112", 2))
    sheet.append(master_row("P2", "Bachelor Two", "COMP1000-112", "COMP1000-111", 2))
    sheet.append(master_row("P3", "Master Three", "DATA5000-111", rule=1, prerequisite=" "))
    workbook.save(path)


class MasterImportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.workbook = self.root / "master.xlsx"
        self.database = self.root / "module_outlines.db"
        write_fixture(self.workbook)
        self.original_db_path = database.DB_PATH
        database.DB_PATH = self.database

    def tearDown(self):
        database.DB_PATH = self.original_db_path
        self.temp.cleanup()

    def test_authoritative_column_mapping_is_complete_and_exact(self):
        self.assertEqual(MASTER_HEADERS, list(import_excel.COLUMN_MAP.values()))

    def test_import_faculty_rules_and_joint_grouping(self):
        result = import_excel.import_data(str(self.workbook), self.database)
        self.assertEqual(1, result["faculties"])
        self.assertEqual(3, result["programmes"])
        self.assertEqual(3, result["classes"])
        self.assertEqual(2, result["outlines"])
        self.assertEqual(1, result["joint_groups"])
        self.assertEqual([], result["warnings"])

        faculties = database.get_faculties()
        self.assertEqual("Faculty of Applied Sciences", faculties[0]["name_en"])
        grouped = database.get_classes_full()
        self.assertEqual(2, len(grouped))
        joint = next(item for item in grouped if item["joint_class"])
        self.assertEqual(["COMP1000-111", "COMP1000-112"], joint["class_codes"])
        self.assertEqual("Bachelor One / Bachelor Two", joint["prog_name_en"])
        self.assertEqual(2, joint["rule_code"])
        self.assertEqual(
            [
                {"class_code": "COMP1000-111", "programme": "Bachelor One"},
                {"class_code": "COMP1000-112", "programme": "Bachelor Two"},
            ],
            joint["programme_class_pairs_en"],
        )

    def test_excel_numeric_integer_values_import_as_rule_codes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(MASTER_HEADERS)
        sheet.append(master_row("P1", "Bachelor One", "COMP1000-111", rule=2))
        sheet.append(master_row("P1", "Bachelor One", "COMP1000-112", rule=2.0))
        workbook.save(self.workbook)

        import_excel.import_data(str(self.workbook), self.database)
        conn = sqlite3.connect(self.database)
        values = [row[0] for row in conn.execute("SELECT rule_code FROM classes ORDER BY class_code")]
        conn.close()
        self.assertEqual([2, 2], values)

    def test_programme_filter_returns_one_complete_joint_outline(self):
        import_excel.import_data(str(self.workbook), self.database)
        programmes = database.get_programmes()
        first_programme = next(item for item in programmes if item["code"] == "P1")
        choices = database.get_classes(programme_id=first_programme["id"])
        self.assertEqual(1, len(choices))
        self.assertEqual("COMP1000-111, COMP1000-112", choices[0]["class_code"])
        generated = database.get_classes_full(programme_id=first_programme["id"])
        self.assertEqual(1, len(generated))
        self.assertEqual(2, generated[0]["joint_member_count"])

    def test_dynamic_years_include_imported_and_future_values(self):
        import_excel.import_data(str(self.workbook), self.database)
        conn = sqlite3.connect(self.database)
        conn.execute("UPDATE classes SET academic_year = '2035/2036' WHERE class_code = 'DATA5000-111'")
        conn.commit()
        conn.close()
        result = database.get_academic_years(date(2026, 9, 8))
        self.assertEqual("2026/2027", result["default"])
        self.assertIn("2029/2030", result["years"])
        self.assertIn("2035/2036", result["years"])

    def test_existing_database_without_rule_is_flagged_for_review(self):
        conn = sqlite3.connect(self.database)
        conn.executescript(
            """
            CREATE TABLE faculties (id INTEGER PRIMARY KEY, code TEXT UNIQUE, name_en TEXT, name_zh TEXT, name_pt TEXT);
            CREATE TABLE programmes (id INTEGER PRIMARY KEY, code TEXT UNIQUE, name_en TEXT, name_zh TEXT, name_pt TEXT, degree_level TEXT, faculty_id INTEGER);
            CREATE TABLE classes (id INTEGER PRIMARY KEY, class_code TEXT UNIQUE, module_code TEXT, programme_id INTEGER);
            INSERT INTO faculties VALUES (1, 'FCA', 'Faculty', '', '');
            INSERT INTO programmes VALUES (1, 'P1', 'Programme', '', '', 'bachelor', 1);
            INSERT INTO classes VALUES (1, 'COMP1000-111', 'COMP1000', 1);
            """
        )
        conn.close()
        with self.assertRaisesRegex(database.RuleMigrationRequiredError, "requires review"):
            database.init_db(self.database, seed=False)
        conn = sqlite3.connect(self.database)
        count = conn.execute("SELECT COUNT(*) FROM classes").fetchone()[0]
        review = conn.execute(
            "SELECT class_code, reason FROM rule_migration_review"
        ).fetchall()
        conn.close()
        self.assertEqual(1, count)
        self.assertEqual("COMP1000-111", review[0][0])
        self.assertIn("blank", review[0][1])

    def test_valid_legacy_marking_rule_is_migrated_without_data_loss(self):
        conn = sqlite3.connect(self.database)
        conn.executescript(
            """
            CREATE TABLE faculties (id INTEGER PRIMARY KEY, code TEXT UNIQUE, name_en TEXT, name_zh TEXT, name_pt TEXT);
            CREATE TABLE programmes (id INTEGER PRIMARY KEY, code TEXT UNIQUE, name_en TEXT, name_zh TEXT, name_pt TEXT, degree_level TEXT, faculty_id INTEGER);
            CREATE TABLE classes (id INTEGER PRIMARY KEY, class_code TEXT UNIQUE, module_code TEXT, marking_rule INTEGER, programme_id INTEGER);
            INSERT INTO faculties VALUES (1, 'FCA', 'Faculty', '', '');
            INSERT INTO programmes VALUES (1, 'P1', 'Programme', '', '', 'bachelor', 1);
            INSERT INTO classes VALUES (1, 'COMP1000-111', 'COMP1000', 2, 1);
            """
        )
        conn.close()
        database.init_db(self.database, seed=False)
        conn = sqlite3.connect(self.database)
        columns = {row[1]: row for row in conn.execute("PRAGMA table_info(classes)")}
        value = conn.execute("SELECT rule_code FROM classes").fetchone()[0]
        table_sql = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='classes'"
        ).fetchone()[0]
        conn.close()
        self.assertEqual(2, value)
        self.assertEqual(1, columns["rule_code"][3])
        self.assertNotIn("marking_rule", columns)
        self.assertIn("CHECK (rule_code IN (1, 2, 3, 4))", table_sql)

    def test_database_constraint_rejects_invalid_rule_codes(self):
        database.init_db(self.database, seed=False)
        conn = sqlite3.connect(self.database)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("INSERT INTO faculties (code) VALUES ('FCA')")
        conn.execute(
            "INSERT INTO programmes (code, degree_level, faculty_id) VALUES ('P1', 'bachelor', 1)"
        )
        for value in (None, 0, 5):
            with self.subTest(value=value), self.assertRaises(sqlite3.IntegrityError):
                conn.execute(
                    "INSERT INTO classes (class_code, module_code, rule_code, programme_id) VALUES (?, ?, ?, 1)",
                    (f"C{value}", "C", value),
                )
        conn.close()

    def test_conflicting_joint_rules_are_reported(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(MASTER_HEADERS)
        sheet.append(master_row("P1", "Bachelor One", "COMP1000-111", "COMP1000-112", 2))
        sheet.append(master_row("P2", "Bachelor Two", "COMP1000-112", "COMP1000-111", 3))
        workbook.save(self.workbook)
        import_excel.import_data(str(self.workbook), self.database)
        choice = database.get_classes()[0]
        self.assertTrue(choice["rule_conflict"])
        self.assertEqual(
            [
                {"class_code": "COMP1000-111", "rule_code": 2},
                {"class_code": "COMP1000-112", "rule_code": 3},
            ],
            choice["rule_conflicts"],
        )
        with self.assertRaisesRegex(JointRuleConflictError, "conflicting Rule"):
            database.get_classes_full()


class RuleNormalizationTests(unittest.TestCase):
    def test_numeric_and_legacy_rule_values(self):
        cases = [
            (1, 1), (2.0, 2), ("3", 3), ("4.0", 4),
            ("Rule 1", 1), ("Rule ONE", 1), ("ONE", 1),
            ("two", 2), ("THREE", 3), ("four", 4), ("一", 1), ("規則四", 4),
        ]
        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(expected, import_excel.normalize_rule(value))

    def test_invalid_rule_values_are_rejected(self):
        for value in (None, "", "  ", "unknown", "Rule FIVE", 0, 5, -1, 2.5, "3.2"):
            with self.subTest(value=value), self.assertRaises(RuleValidationError):
                import_excel.normalize_rule(value)

    def test_invalid_import_is_atomic_and_reports_rows(self):
        for value in (None, "unknown", 2.5, 5):
            with self.subTest(value=value), tempfile.TemporaryDirectory() as directory:
                root = Path(directory)
                workbook_path = root / "invalid.xlsx"
                database_path = root / "existing.db"
                database_path.write_bytes(b"existing database sentinel")
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(MASTER_HEADERS)
                sheet.append(master_row("P1", "Bachelor One", "COMP1000-111", rule=value))
                workbook.save(workbook_path)
                with self.assertRaisesRegex(ValueError, r"row 2 \(COMP1000-111\)"):
                    import_excel.import_data(str(workbook_path), database_path)
                self.assertEqual(b"existing database sentinel", database_path.read_bytes())


if __name__ == "__main__":
    unittest.main()
